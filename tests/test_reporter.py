"""Tests for vpn_collector.reporter (Tasks 7.1–7.6)."""

from __future__ import annotations

import json
import re
from datetime import datetime, timezone
from pathlib import Path

import openpyxl
import pytest

from vpn_collector.collector import DeviceResult
from vpn_collector.config import AppConfig, OutputConfig
from vpn_collector.reporter import write_csv, write_excel, write_json


# ---------------------------------------------------------------------------
# Fixtures / helpers
# ---------------------------------------------------------------------------

_SESSION_A = {
    "username": "alice",
    "public_ip": "1.2.3.4",
    "protocol": "AnyConnect-Parent",
    "bytes_tx": "1024",
    "bytes_rx": "2048",
}

_SESSION_B = {
    "username": "bob",
    "public_ip": "5.6.7.8",
    "protocol": "AnyConnect-Parent",
    "bytes_tx": "512",
    "bytes_rx": "768",
}

_TS = datetime(2024, 6, 1, 12, 0, 0, tzinfo=timezone.utc)


def _make_config(tmp_path: Path) -> AppConfig:
    config = AppConfig()
    config.output = OutputConfig(
        directory=str(tmp_path),
        excel_filename="test-sessions.xlsx",
    )
    return config


def _success(host: str, sessions: list[dict]) -> DeviceResult:
    return DeviceResult(
        host=host,
        success=True,
        sessions=sessions,
        error=None,
        collected_at=_TS,
    )


def _failure(host: str, error: str = "Timed out") -> DeviceResult:
    return DeviceResult(
        host=host,
        success=False,
        sessions=[],
        error=error,
        collected_at=_TS,
    )


# ---------------------------------------------------------------------------
# write_excel — sheet names
# ---------------------------------------------------------------------------

class TestWriteExcelSheetNames:
    def test_summary_sheet_exists(self, tmp_path: Path) -> None:
        config = _make_config(tmp_path)
        results = [_success("fw1", [_SESSION_A])]
        write_excel(results, config)
        wb = openpyxl.load_workbook(tmp_path / "test-sessions.xlsx")
        assert "Summary" in wb.sheetnames

    def test_raw_sheet_exists_with_timestamp(self, tmp_path: Path) -> None:
        config = _make_config(tmp_path)
        results = [_success("fw1", [_SESSION_A])]
        write_excel(results, config)
        wb = openpyxl.load_workbook(tmp_path / "test-sessions.xlsx")
        raw_sheets = [s for s in wb.sheetnames if s.startswith("Raw_")]
        assert len(raw_sheets) == 1

    def test_raw_sheet_name_format(self, tmp_path: Path) -> None:
        config = _make_config(tmp_path)
        results = [_success("fw1", [_SESSION_A])]
        write_excel(results, config)
        wb = openpyxl.load_workbook(tmp_path / "test-sessions.xlsx")
        raw_sheets = [s for s in wb.sheetnames if s.startswith("Raw_")]
        assert re.fullmatch(r"Raw_\d{8}_\d{6}", raw_sheets[0])

    def test_exactly_two_sheets_on_first_write(self, tmp_path: Path) -> None:
        config = _make_config(tmp_path)
        results = [_success("fw1", [_SESSION_A])]
        write_excel(results, config)
        wb = openpyxl.load_workbook(tmp_path / "test-sessions.xlsx")
        assert len(wb.sheetnames) == 2


# ---------------------------------------------------------------------------
# write_excel — raw sheet content
# ---------------------------------------------------------------------------

class TestWriteExcelRawSheet:
    def test_raw_sheet_has_header_row(self, tmp_path: Path) -> None:
        config = _make_config(tmp_path)
        results = [_success("fw1", [_SESSION_A])]
        write_excel(results, config)
        wb = openpyxl.load_workbook(tmp_path / "test-sessions.xlsx")
        raw_sheet = next(wb[s] for s in wb.sheetnames if s.startswith("Raw_"))
        header = [cell.value for cell in raw_sheet[1]]
        assert header == list(_SESSION_A.keys())

    def test_raw_sheet_data_rows_count(self, tmp_path: Path) -> None:
        config = _make_config(tmp_path)
        results = [_success("fw1", [_SESSION_A, _SESSION_B])]
        write_excel(results, config)
        wb = openpyxl.load_workbook(tmp_path / "test-sessions.xlsx")
        raw_sheet = next(wb[s] for s in wb.sheetnames if s.startswith("Raw_"))
        # Row 1 = header; rows 2+ = data
        data_rows = list(raw_sheet.iter_rows(min_row=2, values_only=True))
        assert len(data_rows) == 2

    def test_raw_sheet_values_match_session(self, tmp_path: Path) -> None:
        config = _make_config(tmp_path)
        results = [_success("fw1", [_SESSION_A])]
        write_excel(results, config)
        wb = openpyxl.load_workbook(tmp_path / "test-sessions.xlsx")
        raw_sheet = next(wb[s] for s in wb.sheetnames if s.startswith("Raw_"))
        header = [cell.value for cell in raw_sheet[1]]
        data_row = [cell.value for cell in raw_sheet[2]]
        record = dict(zip(header, data_row))
        assert record["username"] == "alice"
        assert record["public_ip"] == "1.2.3.4"


# ---------------------------------------------------------------------------
# write_excel — Summary sheet content
# ---------------------------------------------------------------------------

class TestWriteExcelSummarySheet:
    def test_summary_contains_run_timestamp(self, tmp_path: Path) -> None:
        config = _make_config(tmp_path)
        results = [_success("fw1", [_SESSION_A])]
        write_excel(results, config)
        wb = openpyxl.load_workbook(tmp_path / "test-sessions.xlsx")
        ws = wb["Summary"]
        # iter_rows(values_only=True) yields tuples of plain values
        flat = [v for row in ws.iter_rows(values_only=True) for v in row if v is not None]
        assert "Run Timestamp" in flat

    def test_summary_device_table_header(self, tmp_path: Path) -> None:
        config = _make_config(tmp_path)
        results = [_success("fw1", [_SESSION_A])]
        write_excel(results, config)
        wb = openpyxl.load_workbook(tmp_path / "test-sessions.xlsx")
        ws = wb["Summary"]
        header_found = False
        for row in ws.iter_rows(values_only=True):
            if "Host" in row:
                header_found = True
                break
        assert header_found

    def test_summary_lists_device_host(self, tmp_path: Path) -> None:
        config = _make_config(tmp_path)
        results = [_success("fw1.example.com", [_SESSION_A])]
        write_excel(results, config)
        wb = openpyxl.load_workbook(tmp_path / "test-sessions.xlsx")
        ws = wb["Summary"]
        all_values = [cell for row in ws.iter_rows(values_only=True) for cell in row]
        assert "fw1.example.com" in all_values


# ---------------------------------------------------------------------------
# write_excel — append (existing workbook)
# ---------------------------------------------------------------------------

class TestWriteExcelAppend:
    def test_append_adds_new_raw_sheet(self, tmp_path: Path) -> None:
        """Second call must add a second Raw_ sheet without removing the first."""
        config = _make_config(tmp_path)
        results = [_success("fw1", [_SESSION_A])]
        write_excel(results, config)
        write_excel(results, config)
        wb = openpyxl.load_workbook(tmp_path / "test-sessions.xlsx")
        raw_sheets = [s for s in wb.sheetnames if s.startswith("Raw_")]
        assert len(raw_sheets) == 2

    def test_append_preserves_first_raw_sheet(self, tmp_path: Path) -> None:
        config = _make_config(tmp_path)
        results = [_success("fw1", [_SESSION_A])]
        write_excel(results, config)
        wb_first = openpyxl.load_workbook(tmp_path / "test-sessions.xlsx")
        first_raw = next(s for s in wb_first.sheetnames if s.startswith("Raw_"))

        write_excel(results, config)
        wb_second = openpyxl.load_workbook(tmp_path / "test-sessions.xlsx")
        assert first_raw in wb_second.sheetnames

    def test_append_summary_sheet_is_replaced_not_duplicated(self, tmp_path: Path) -> None:
        config = _make_config(tmp_path)
        results = [_success("fw1", [_SESSION_A])]
        write_excel(results, config)
        write_excel(results, config)
        wb = openpyxl.load_workbook(tmp_path / "test-sessions.xlsx")
        assert wb.sheetnames.count("Summary") == 1

    def test_append_total_sheet_count_grows_by_one(self, tmp_path: Path) -> None:
        config = _make_config(tmp_path)
        results = [_success("fw1", [_SESSION_A])]
        write_excel(results, config)
        wb1 = openpyxl.load_workbook(tmp_path / "test-sessions.xlsx")
        count_after_first = len(wb1.sheetnames)

        write_excel(results, config)
        wb2 = openpyxl.load_workbook(tmp_path / "test-sessions.xlsx")
        assert len(wb2.sheetnames) == count_after_first + 1


# ---------------------------------------------------------------------------
# write_csv
# ---------------------------------------------------------------------------

class TestWriteCsv:
    def _csv_path(self, tmp_path: Path) -> Path:
        files = list(tmp_path.glob("anyconnect-sessions-*.csv"))
        assert len(files) == 1, f"Expected 1 CSV, found {len(files)}"
        return files[0]

    def test_csv_file_created(self, tmp_path: Path) -> None:
        config = _make_config(tmp_path)
        write_csv([_success("fw1", [_SESSION_A])], config)
        files = list(tmp_path.glob("anyconnect-sessions-*.csv"))
        assert len(files) == 1

    def test_csv_filename_pattern(self, tmp_path: Path) -> None:
        config = _make_config(tmp_path)
        write_csv([_success("fw1", [_SESSION_A])], config)
        files = list(tmp_path.glob("anyconnect-sessions-*.csv"))
        assert re.fullmatch(r"anyconnect-sessions-\d{8}_\d{6}\.csv", files[0].name)

    def test_csv_header_row_present(self, tmp_path: Path) -> None:
        config = _make_config(tmp_path)
        write_csv([_success("fw1", [_SESSION_A])], config)
        csv_file = self._csv_path(tmp_path)
        lines = csv_file.read_text(encoding="utf-8").splitlines()
        header = lines[0].split(",")
        assert header == list(_SESSION_A.keys())

    def test_csv_correct_number_of_data_rows(self, tmp_path: Path) -> None:
        config = _make_config(tmp_path)
        results = [_success("fw1", [_SESSION_A, _SESSION_B])]
        write_csv(results, config)
        csv_file = self._csv_path(tmp_path)
        lines = csv_file.read_text(encoding="utf-8").splitlines()
        # 1 header + 2 data rows
        assert len(lines) == 3

    def test_csv_multiple_devices(self, tmp_path: Path) -> None:
        config = _make_config(tmp_path)
        results = [
            _success("fw1", [_SESSION_A]),
            _success("fw2", [_SESSION_B]),
        ]
        write_csv(results, config)
        csv_file = self._csv_path(tmp_path)
        lines = csv_file.read_text(encoding="utf-8").splitlines()
        # 1 header + 2 data rows across two devices
        assert len(lines) == 3

    def test_csv_values_present(self, tmp_path: Path) -> None:
        config = _make_config(tmp_path)
        write_csv([_success("fw1", [_SESSION_A])], config)
        csv_file = self._csv_path(tmp_path)
        content = csv_file.read_text(encoding="utf-8")
        assert "alice" in content
        assert "1.2.3.4" in content


# ---------------------------------------------------------------------------
# write_json
# ---------------------------------------------------------------------------

class TestWriteJson:
    def _json_path(self, tmp_path: Path) -> Path:
        files = list(tmp_path.glob("anyconnect-sessions-*.json"))
        assert len(files) == 1, f"Expected 1 JSON, found {len(files)}"
        return files[0]

    def test_json_file_created(self, tmp_path: Path) -> None:
        config = _make_config(tmp_path)
        write_json([_success("fw1", [_SESSION_A])], config)
        files = list(tmp_path.glob("anyconnect-sessions-*.json"))
        assert len(files) == 1

    def test_json_filename_pattern(self, tmp_path: Path) -> None:
        config = _make_config(tmp_path)
        write_json([_success("fw1", [_SESSION_A])], config)
        files = list(tmp_path.glob("anyconnect-sessions-*.json"))
        assert re.fullmatch(r"anyconnect-sessions-\d{8}_\d{6}\.json", files[0].name)

    def test_json_is_array(self, tmp_path: Path) -> None:
        config = _make_config(tmp_path)
        write_json([_success("fw1", [_SESSION_A])], config)
        data = json.loads(self._json_path(tmp_path).read_text(encoding="utf-8"))
        assert isinstance(data, list)

    def test_json_array_length_matches_devices(self, tmp_path: Path) -> None:
        config = _make_config(tmp_path)
        results = [_success("fw1", [_SESSION_A]), _success("fw2", [_SESSION_B])]
        write_json(results, config)
        data = json.loads(self._json_path(tmp_path).read_text(encoding="utf-8"))
        assert len(data) == 2

    def test_json_element_has_required_keys(self, tmp_path: Path) -> None:
        config = _make_config(tmp_path)
        write_json([_success("fw1", [_SESSION_A])], config)
        data = json.loads(self._json_path(tmp_path).read_text(encoding="utf-8"))
        entry = data[0]
        assert "host" in entry
        assert "success" in entry
        assert "collected_at" in entry
        assert "session_count" in entry
        assert "sessions" in entry

    def test_json_collected_at_is_iso_8601(self, tmp_path: Path) -> None:
        config = _make_config(tmp_path)
        write_json([_success("fw1", [_SESSION_A])], config)
        data = json.loads(self._json_path(tmp_path).read_text(encoding="utf-8"))
        collected_at = data[0]["collected_at"]
        # Should be parseable as ISO 8601
        dt = datetime.fromisoformat(collected_at)
        assert dt is not None

    def test_json_sessions_list(self, tmp_path: Path) -> None:
        config = _make_config(tmp_path)
        write_json([_success("fw1", [_SESSION_A])], config)
        data = json.loads(self._json_path(tmp_path).read_text(encoding="utf-8"))
        assert isinstance(data[0]["sessions"], list)
        assert data[0]["sessions"][0]["username"] == "alice"

    def test_json_failed_device_success_false(self, tmp_path: Path) -> None:
        config = _make_config(tmp_path)
        results = [
            _success("fw1", [_SESSION_A]),
            _failure("fw2", "Connection refused"),
        ]
        write_json(results, config)
        data = json.loads(self._json_path(tmp_path).read_text(encoding="utf-8"))
        failed = next(e for e in data if e["host"] == "fw2")
        assert failed["success"] is False

    def test_json_failed_device_has_error_field(self, tmp_path: Path) -> None:
        config = _make_config(tmp_path)
        results = [
            _success("fw1", [_SESSION_A]),
            _failure("fw2", "Connection refused"),
        ]
        write_json(results, config)
        data = json.loads(self._json_path(tmp_path).read_text(encoding="utf-8"))
        failed = next(e for e in data if e["host"] == "fw2")
        assert "error" in failed
        assert failed["error"] == "Connection refused"

    def test_json_failed_device_included_in_array(self, tmp_path: Path) -> None:
        """Failed devices must appear in the JSON output, not be silently dropped."""
        config = _make_config(tmp_path)
        results = [
            _success("fw1", [_SESSION_A]),
            _failure("fw2"),
        ]
        write_json(results, config)
        data = json.loads(self._json_path(tmp_path).read_text(encoding="utf-8"))
        hosts = [e["host"] for e in data]
        assert "fw2" in hosts


# ---------------------------------------------------------------------------
# Empty results guard (Task 7.5)
# ---------------------------------------------------------------------------

class TestEmptyResultsGuard:
    def test_write_excel_no_file_when_empty(self, tmp_path: Path) -> None:
        config = _make_config(tmp_path)
        write_excel([], config)
        assert not (tmp_path / "test-sessions.xlsx").exists()

    def test_write_excel_no_file_when_all_failed(self, tmp_path: Path) -> None:
        config = _make_config(tmp_path)
        results = [_failure("fw1"), _failure("fw2")]
        write_excel(results, config)
        assert not (tmp_path / "test-sessions.xlsx").exists()

    def test_write_csv_no_file_when_empty(self, tmp_path: Path) -> None:
        config = _make_config(tmp_path)
        write_csv([], config)
        csv_files = list(tmp_path.glob("anyconnect-sessions-*.csv"))
        assert csv_files == []

    def test_write_csv_no_file_when_all_failed(self, tmp_path: Path) -> None:
        config = _make_config(tmp_path)
        results = [_failure("fw1"), _failure("fw2")]
        write_csv(results, config)
        csv_files = list(tmp_path.glob("anyconnect-sessions-*.csv"))
        assert csv_files == []

    def test_write_json_no_file_when_empty(self, tmp_path: Path) -> None:
        config = _make_config(tmp_path)
        write_json([], config)
        json_files = list(tmp_path.glob("anyconnect-sessions-*.json"))
        assert json_files == []

    def test_write_json_no_file_when_all_failed(self, tmp_path: Path) -> None:
        config = _make_config(tmp_path)
        results = [_failure("fw1"), _failure("fw2")]
        write_json(results, config)
        json_files = list(tmp_path.glob("anyconnect-sessions-*.json"))
        assert json_files == []

    def test_write_excel_logs_warning_when_empty(
        self, tmp_path: Path, caplog: pytest.LogCaptureFixture
    ) -> None:
        import logging
        config = _make_config(tmp_path)
        pkg_logger = logging.getLogger("vpn_collector")
        original_propagate = pkg_logger.propagate
        try:
            pkg_logger.propagate = True
            with caplog.at_level(logging.WARNING, logger="vpn_collector.reporter"):
                write_excel([], config)
        finally:
            pkg_logger.propagate = original_propagate
        assert "No session data to write" in caplog.text

    def test_write_csv_logs_warning_when_empty(
        self, tmp_path: Path, caplog: pytest.LogCaptureFixture
    ) -> None:
        import logging
        config = _make_config(tmp_path)
        pkg_logger = logging.getLogger("vpn_collector")
        original_propagate = pkg_logger.propagate
        try:
            pkg_logger.propagate = True
            with caplog.at_level(logging.WARNING, logger="vpn_collector.reporter"):
                write_csv([], config)
        finally:
            pkg_logger.propagate = original_propagate
        assert "No session data to write" in caplog.text

    def test_write_json_logs_warning_when_empty(
        self, tmp_path: Path, caplog: pytest.LogCaptureFixture
    ) -> None:
        import logging
        config = _make_config(tmp_path)
        pkg_logger = logging.getLogger("vpn_collector")
        original_propagate = pkg_logger.propagate
        try:
            pkg_logger.propagate = True
            with caplog.at_level(logging.WARNING, logger="vpn_collector.reporter"):
                write_json([], config)
        finally:
            pkg_logger.propagate = original_propagate
        assert "No session data to write" in caplog.text
