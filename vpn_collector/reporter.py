"""Output writers for the ASA VPN session collector.

Provides three write functions:
  - write_excel  — appends to (or creates) an .xlsx workbook
  - write_csv    — writes a flat CSV of all sessions
  - write_json   — serialises the full list[DeviceResult] to JSON
"""

from __future__ import annotations

import csv
import json
import logging
import os
from datetime import datetime, timezone
from pathlib import Path

import openpyxl
from openpyxl.styles import Font

from vpn_collector.collector import DeviceResult
from vpn_collector.config import AppConfig

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Internal helpers
# ---------------------------------------------------------------------------

_NO_DATA_MSG = "No session data to write — skipping output."


def _has_sessions(results: list[DeviceResult]) -> bool:
    """Return True if there is at least one session across all results."""
    return sum(len(r.sessions) for r in results) > 0


def _all_session_fields(results: list[DeviceResult]) -> list[str]:
    """Collect an ordered, deduplicated list of all session dict keys."""
    seen: dict[str, None] = {}
    for result in results:
        for session in result.sessions:
            for key in session:
                seen[key] = None
    return list(seen.keys())


def _run_timestamp() -> str:
    """Return a timestamp string suitable for sheet/file names: YYYYMMDD_HHMMSS."""
    return datetime.now().strftime("%Y%m%d_%H%M%S")


def _ensure_dir(directory: str) -> Path:
    """Create *directory* if it does not exist and return it as a Path."""
    path = Path(directory)
    path.mkdir(parents=True, exist_ok=True)
    return path


# ---------------------------------------------------------------------------
# Excel writer (Task 7.1 / 7.2)
# ---------------------------------------------------------------------------

def write_excel(results: list[DeviceResult], config: AppConfig) -> None:
    """Write (or append to) an Excel workbook with Summary and Raw sheets.

    Parameters
    ----------
    results:
        Collection results, one entry per device.
    config:
        Application configuration; ``config.output.directory`` and
        ``config.output.excel_filename`` control the output path.
    """
    if not _has_sessions(results):
        logger.warning(_NO_DATA_MSG)
        return

    out_dir = _ensure_dir(config.output.directory)
    xlsx_path = out_dir / config.output.excel_filename

    # ------------------------------------------------------------------
    # Load existing workbook or create a new one
    # ------------------------------------------------------------------
    if xlsx_path.exists():
        wb = openpyxl.load_workbook(xlsx_path)
    else:
        wb = openpyxl.Workbook()
        # openpyxl creates a default "Sheet" — remove it
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

    ts = _run_timestamp()

    # ------------------------------------------------------------------
    # Summary sheet — always replaced with the latest run data
    # ------------------------------------------------------------------
    if "Summary" in wb.sheetnames:
        del wb["Summary"]

    # Insert Summary as the first sheet
    ws_summary = wb.create_sheet("Summary", 0)

    run_ts = datetime.now(timezone.utc).isoformat()
    total_devices = len(results)
    successful_devices = sum(1 for r in results if r.success)
    total_sessions = sum(len(r.sessions) for r in results)

    # Run metadata block
    ws_summary.append(["Run Timestamp", run_ts])
    ws_summary.append(["Total Devices", total_devices])
    ws_summary.append(["Successful Devices", successful_devices])
    ws_summary.append(["Total Sessions", total_sessions])
    ws_summary.append([])  # blank separator

    # Per-device table header
    header = ["Host", "Sessions", "Status"]
    ws_summary.append(header)
    for cell in ws_summary[ws_summary.max_row]:
        cell.font = Font(bold=True)

    for result in results:
        status = "success" if result.success else f"failed: {result.error}"
        ws_summary.append([result.host, len(result.sessions), status])

    # ------------------------------------------------------------------
    # Raw sheet — one per run, named Raw_YYYYMMDD_HHMMSS
    # ------------------------------------------------------------------
    raw_sheet_name = f"Raw_{ts}"
    ws_raw = wb.create_sheet(raw_sheet_name)

    fields = _all_session_fields(results)
    ws_raw.append(fields)
    for cell in ws_raw[1]:
        cell.font = Font(bold=True)

    for result in results:
        for session in result.sessions:
            row = [session.get(f, "") for f in fields]
            ws_raw.append(row)

    wb.save(xlsx_path)
    logger.info("Excel written: %s (raw sheet: %s)", xlsx_path, raw_sheet_name)


# ---------------------------------------------------------------------------
# CSV writer (Task 7.3)
# ---------------------------------------------------------------------------

def write_csv(results: list[DeviceResult], config: AppConfig) -> None:
    """Write all sessions to a flat CSV file.

    Parameters
    ----------
    results:
        Collection results, one entry per device.
    config:
        Application configuration; ``config.output.directory`` controls
        where the file is written.
    """
    if not _has_sessions(results):
        logger.warning(_NO_DATA_MSG)
        return

    out_dir = _ensure_dir(config.output.directory)
    ts = _run_timestamp()
    csv_path = out_dir / f"anyconnect-sessions-{ts}.csv"

    fields = _all_session_fields(results)

    with csv_path.open("w", newline="", encoding="utf-8") as fh:
        writer = csv.DictWriter(fh, fieldnames=fields, extrasaction="ignore")
        writer.writeheader()
        for result in results:
            for session in result.sessions:
                writer.writerow(session)

    logger.info("CSV written: %s", csv_path)


# ---------------------------------------------------------------------------
# JSON writer (Task 7.4)
# ---------------------------------------------------------------------------

def write_json(results: list[DeviceResult], config: AppConfig) -> None:
    """Write all results to a JSON file.

    Parameters
    ----------
    results:
        Collection results, one entry per device.
    config:
        Application configuration; ``config.output.directory`` controls
        where the file is written.
    """
    if not _has_sessions(results):
        logger.warning(_NO_DATA_MSG)
        return

    out_dir = _ensure_dir(config.output.directory)
    ts = _run_timestamp()
    json_path = out_dir / f"anyconnect-sessions-{ts}.json"

    payload: list[dict] = []
    for result in results:
        entry: dict = {
            "host": result.host,
            "success": result.success,
            "collected_at": result.collected_at.isoformat(),
            "session_count": len(result.sessions),
            "sessions": result.sessions,
        }
        if not result.success and result.error is not None:
            entry["error"] = result.error
        payload.append(entry)

    with json_path.open("w", encoding="utf-8") as fh:
        json.dump(payload, fh, indent=2, default=str)

    logger.info("JSON written: %s", json_path)
